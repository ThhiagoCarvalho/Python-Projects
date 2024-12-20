from openpyxl import Workbook, load_workbook
import pandas as pd

situcao = "REPROVADO"
alunos_lista = []


def validar_notas(mensagem_input):
    while True:
        valor = input(mensagem_input)
        try:
            while (float(valor) < 0.0 or float(valor) > 10.0):
                valor = input(f"DIGITE CORRETAMENTE a nota do aluno {nome} entre 0 e 10: ")
            return round(float(valor), 1)
        except ValueError:
            print("DIGITE CORRETAMENTE a nota do aluno!")


qtd = input("Deseja cadastrar quantos alunos na tabela Excel?: ")
while (not qtd.isnumeric() or int(qtd) <= 0):
    qtd = input("DIGITE CORRETAMENTE! deseja cadastrar quantos alunos na tabela Excel?: ")

for cont in range(int(qtd)):
    media = 0
    print("=-" * 60)

    rm = input(f"Digite o RM do {cont + 1}º aluno: ")
    while (not rm.isnumeric() or len(rm) != 6):
        rm = input(f"DIGITE CORRETAMENTE o RM do {cont + 1}º aluno:")

    nome = input(f"Digite o nome do {cont + 1}º aluno: ")
    while (nome == '' or nome.isnumeric()):
        nome = input(f"DIGITE CORRETAMENTE o nome do {cont + 1}º aluno!: ")

    notas_pvb = validar_notas(f"Digite a nota do aluno {nome} na matéria PVB: ")
    notas_pooi = validar_notas(f"Digite a nota do aluno {nome} na na matéria POOI: ")
    notas_paw = validar_notas(f"Digite a nota do aluno {nome} na matéria PAW: ")
    notas_bd = validar_notas(f"Digite a nota do aluno {nome} na matéria BD: ")




    media = (notas_bd + notas_paw + notas_pooi + notas_pvb) / 4
    media = round(float(media), 1)
    if (media > 3.75 and media <= 5.9):
        situcao = "EXAME FINAL"
    elif (media >= 6):
        situcao = "APROVADO"

    aluno = [rm, nome, notas_pvb, notas_paw, notas_bd, notas_pooi, media, situcao]

    alunos_lista.append(aluno)

caminho_arquivo = "C:/projeto_4bim/NOTASFINAISALUNOS.xlsx"

print("=-" * 60)
print()
print("1- Adicionar os dados em uma planilha nova \n2- Adicionar os dados em planilha existente\n")
resp = input("Digite sua escolha (1 ou 2): ")

while (resp != "1" and resp != "2"):
    print("=-" * 60)
    print("\n1- Adicionar os dados em uma planilha nova \n2- Adicionar os dados em planilha existente\n")
    resp = input("DIGITE CORRETAMENTE: ")

if resp == "1":

    wb = Workbook()
    ws = wb.active
    ws.title = "NOTASFINAISALUNOS"

    # Adiciona cabeçalho
    ws.append(["MATRICULA", "ALUNOS", "PVB", "PAW", "BD", "POOI", "MEDIA"])

    for aluno in alunos_lista:
        ws.append(aluno[0:7])

    wb.save(caminho_arquivo)

elif resp == "2":
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb["NOTASFINAISALUNOS"]

        for aluno in alunos_lista:
            ws.append(aluno)

        wb.save(caminho_arquivo)

    except FileNotFoundError:
        print("Erro: Arquivo não encontrado. Certifique-se de que o caminho está correto.")

print("=-" * 60)

resp_html = input("\ndeseja criar uma pagina html para demonstrar as notas? (s/n): ")
while (resp_html != "s" and resp_html != "n"):
    resp_html = input("DIGITE CORRETAMENTE (s/n): ")

if (resp_html == 's'):
    for aluno in alunos_lista:
        nomeArquivo = aluno[0] + ".html"
        arquivo = open(f'C:/projeto_4bim/{nomeArquivo}', 'w',
                       encoding='utf-8')
        body = ''

        background = "rgb(29, 99, 192)"
        if (aluno[6] > 3.75 and aluno[6] <= 5.9):
            background = "yellow"
        elif (aluno[6] < 3.75):
            background = "red"

        html = f"""
        <!DOCTYPE html>
        <html lang="UTF-8">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Notas Finais</title>

            <style>
            body{{    
                text-align: center;
                display: block;
                justify-content: center;
                align-items: center;
            }}
            p{{margin-bottom : 40px}}

            table{{
                margin: auto;
                background-color: {background};
                border-collapse:collapse;

            }}
            tr{{
                border: 1px solid;
            }}

            #red {{
                background-color: red
            }}
                #situacao {{
                background-color: {background};
            }}
            </style>
        </head>
        <body>
        <p> MATRICULA  : {aluno[0]}</p>
        <p>ALUNO: <span id="red">{aluno[1]}</span></p>


        <table>
        <theads>
            <th>PVB</th>
            <th>PAW</th>
            <th>BD</th>
            <th>POOI</th>
            <th>MEDIA</th>
            </theads>
            <tr>

            <td> {aluno[2]} </td>
            <td> {aluno[3]} </td>
            <td> {aluno[4]}</td>
            <td> {aluno[5]}</td>
            <td> {aluno[6]}</td>
            </tr>
        </table>

        <p> SITUACAO FINAL  <span id="situacao">{aluno[-1]}</span> </p>

        </body>

        </html>"""

        arquivo.write(html)
        print(f"Arquivo '{nomeArquivo}' criado com sucesso!")
        arquivo.close()
